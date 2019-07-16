import logging

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from flask import abort

from excel_lib.actions import (
    merge,
    unmerge,
)
from excel_lib.constants import add_units
from excel_lib.table import Table
from service_app.domain.data_master import get_materials_data, get_next_abbreviation_pack, get_merge_cells
from service_app.constants import DEFAULT_SHEET_NAME
from service_app.domain.configs import Config


def add_column_with_prices(table, table_prices, symbol):
    rows_number = table.rows_number
    new_table = table
    index = True
    while index:
        materials_row = new_table.get_row_with_parameters()
        table_prices_abbreviation = table_prices.get_column(1)
        materials, column_number, index = get_next_abbreviation_pack(
            index, materials_row, table_prices_abbreviation, symbol
        )
        index = index
        if materials and column_number:
            materials_data = get_materials_data(materials, table_prices, rows_number, column_number)
            new_table.add_multiple_columns(column_number + 2, materials_data)

    return new_table


def copy_worksheet(book, sheet_name, new_sheet_name):
    try:
        sheet = book.get_sheet_by_name(new_sheet_name)
        book.remove(sheet)
        logging.info(f"Sheet {sheet_name} remove")
    except KeyError:
        logging.info(f"Sheet {sheet_name} not exist")
    finally:
        sheet = book.get_sheet_by_name(sheet_name)
        book.copy_worksheet(sheet)
        sheet = book.get_sheet_by_name(new_sheet_name)
        logging.info('Create new sheet')
    return sheet


def start_modify(config):
    rows = config.header_rows
    try:
        book = openpyxl.load_workbook(config.file_path)
    except FileNotFoundError:
        abort(404, f"Excel file not found in {config.file_path}")
    config.make_new_worksheet()
    sheet = copy_worksheet(book, config.sheet_name, config.new_sheet_name)
    unmerge(sheet)
    try:
        sheet_prices = book.get_sheet_by_name(config.prices_sheet_name)
    except KeyError:
        abort(404, f"Sheet {config.prices_sheet_name} not found")
    table = Table(sheet, config.row_with_params_number)
    for row in rows:
        constants_data = table.get_row_values(row)
        add_units(constants_data)
    table_prices = Table(sheet=sheet_prices)
    symbol = config.separator_symbol
    if not symbol:
        symbol = ';'
    logging.info(f"Start with separation symbol {symbol} in file {config.file_path}")
    new_table = add_column_with_prices(table, table_prices, symbol)
    logging.info(f"Added new columns in file {config.file_path}")
    new_table.write_table(sheet)
    logging.info(f"New tables wrote in file {config.file_path}")
    merge_rows_numbers = [1, 2]
    for row_number in merge_rows_numbers:
        row_data = new_table.get_row_values(row_number)
        merge_cells = get_merge_cells(row_data, row_number)
        merge(sheet, merge_cells)

    try:
        book.save(config.file_path)
        logging.info(f"Excell save {config.file_path}")
    except Exception:
        abort(500, "Modify excel file could not be saved")
    return new_table.preview()


def get_table_preview(file_path, sheet_name=None):
    try:
        book = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        abort(404, f"File {file_path} not found")
    except InvalidFileException:
        abort(400, f"File {file_path} has wrong format. Supported formats are: .xlsx,.xlsm,.xltx,.xltm")
    if not sheet_name:
        sheet = book.get_sheet_by_name(DEFAULT_SHEET_NAME)
        logging.info(f"Generating table preview for sheet: {DEFAULT_SHEET_NAME}")
    else:
        sheet = book.get_sheet_by_name(sheet_name)
        logging.info(f"Generating table preview for sheet: {sheet_name}")
    table = Table(sheet)
    return table.preview()


def modify_table(file_path, sheet_name, sheet_price_name, row_number, separate_symbol):
    config_data = {
        "file_path": file_path,
        "sheet_name": sheet_name,
        "prices_sheet_name": sheet_price_name,
        "row_with_params_number": int(row_number),
        "separator_symbol": separate_symbol
    }
    logging.info(f"Config for modify: {config_data}")
    config = Config().load_configs(config_data)
    new_table = start_modify(config)
    return new_table
