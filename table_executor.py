import openpyxl
import logging

from data_master import get_materials_data, get_next_abbreviation_pack, get_merge_cells
from excell_lib.actions import (
    merge,
    unmerge,
)
from excell_lib.constants import add_units
from excell_lib.table import Table


def add_column_with_prices(table, table_prices, symbol):
    rows_number = table.rows_number
    new_table = table
    index = True
    while index:
        materials_row = new_table.get_row_with_parameters()
        table_prices_abbreviation = table_prices.get_column(1)
        materials, column_number, index = get_next_abbreviation_pack(index, materials_row, table_prices_abbreviation, symbol)
        index = index
        if materials and column_number:
            materials_data = get_materials_data(materials, table_prices, rows_number, column_number)
            new_table.add_multiple_columns(column_number + 2, materials_data)

    return new_table


def copy_worksheet(book, sheet_name, new_sheet_name):
    try:
        sheet = book.get_sheet_by_name(new_sheet_name)
        book.remove(sheet)
    except KeyError as err:
        logging.info(str(err))
        sheet = book.get_sheet_by_name(sheet_name)
        book.copy_worksheet(sheet)
        sheet = book.get_sheet_by_name(new_sheet_name)
        logging.info('Create new sheet')
    return sheet


def start(config):
    logging.info('Start excel program')
    rows = config.header_rows
    book = openpyxl.load_workbook(config.file_path)
    sheet = copy_worksheet(book, config.sheet_name, config.new_sheet_name)
    unmerge(sheet)
    sheet_prices = book.get_sheet_by_name(config.prices_sheet_name)
    table = Table(sheet, config.row_with_params_number)
    for row in rows:
        constants_data = table.get_row_values(row)
        add_units(constants_data)
    logging.info('Create main table')
    table_prices = Table(sheet=sheet_prices)
    logging.info('Create prices table')
    symbol = config.separator_symbol
    if not symbol:
        symbol = ';'
    new_table = add_column_with_prices(table, table_prices, symbol)
    logging.info('New columns created')
    new_table.write_table(sheet)
    logging.info('New tables wrote')
    merge_rows_numbers = [1, 2]
    for row_number in merge_rows_numbers:
        row_data = new_table.get_row_values(row_number)
        merge_cells = get_merge_cells(row_data, row_number)
        merge(sheet, merge_cells)
    logging.info('Automerge complete')

    book.save(config.file_path)
    logging.info('Excell save')
